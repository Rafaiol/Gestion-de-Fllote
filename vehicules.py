import tkinter as tk
import customtkinter as ctk
from tkinter import messagebox
from tkinter import ttk
import pyodbc
from PIL import Image, ImageTk
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import datetime

ctk.set_appearance_mode("dark")  # Modes: "dark", "light", "system"
ctk.set_default_color_theme("blue")  # Themes: "blue", "green", "dark-blue"

class MainPage(ctk.CTkFrame):
    def load_icon(self, path, size):
         """Load and resize an icon."""
         img =  Image.open(path)
         img_resized = img.resize(size, Image.Resampling.LANCZOS)
         return ImageTk.PhotoImage(img_resized) 
    def __init__(self, master=None,user_role=None, **kwargs):
        super().__init__(master, **kwargs)
        self.user_role = user_role
        self.pack(fill="both", expand=True)
        
        self.existing_button_frames = {}
        
        self.add_icon = self.load_icon("Image_Assets/Addition.png", size=(25, 25))
        self.filter_icon = self.load_icon("Image_Assets/sort.png", size=(25, 25))
        self.refresh_icon = self.load_icon("Image_Assets/Refresh.png", size=(25, 25))
        self.search_icon = self.load_icon("Image_Assets/search.png", size=(25, 25))
        self.supprimer_icon = self.load_icon("Image_Assets/trash.png", size=(20, 20))
        self.Update_icon = self.load_icon("Image_Assets/edit.png", size=(20, 20))
        self.Inspect_icon = self.load_icon("Image_Assets/search1.png", size=(20, 20))
        
        self.page_frame = ctk.CTkFrame(self, corner_radius=10,fg_color="#050505")
        self.page_frame.pack(side="right", fill="both", expand=True, padx=10, pady=10)
        self.page_frame.grid_columnconfigure(0, weight=1)
        self.page_frame.grid_rowconfigure(1, weight=1)
       
        
        
        
        self.buttons_frame = ctk.CTkFrame(self.page_frame, corner_radius=20,height=80,fg_color="transparent")
        self.buttons_frame.grid(row=0,column=0,columnspan=2, padx=10,pady=10,sticky="ew")
        self.buttons_frame.grid_columnconfigure(0, weight=1)
        self.buttons_frame.grid_rowconfigure(0, weight=1)
        
        self.right_buttons_frame = ctk.CTkFrame(self.page_frame,fg_color="transparent" ,corner_radius=0,width=100)
        self.right_buttons_frame.grid(row=1,column=1,padx=(3,3) ,pady=10,sticky="nsew")
  
        self.actions_label = ctk.CTkLabel(self.right_buttons_frame, text="Actions", font=("Poppins", 14, "bold"))
        self.actions_label.pack(fill="x")

    
    
        
        
        
# Database connection helper function
        def get_connection(): 
         try:
          return pyodbc.connect(
            'DRIVER={ODBC Driver 17 for SQL Server};'
            'server=WIN-1QQFTLB3RC8;'
            'Database=gestion_de_fllote;'
            'Trusted_Connection=yes'
        )
         except pyodbc.Error as e:
          messagebox.showerror("Database Error", f"Failed to connect: {str(e)}")
          return None
      

        
        def start_add_mode(tree, tab):
            
                           
            # Create popup window
            popup = ctk.CTkToplevel()
            popup.title("Add New Record")
            popup.overrideredirect(True)
            popup.geometry("700x450")
            popup.resizable(False, False)
            popup.transient(tree.winfo_toplevel())  # Make popup transient to main window
            popup.grab_set()  # Make popup modal
            popup.focus_set()
        
                    # Create main frame with padding
            main_frame = ctk.CTkFrame(popup)
            main_frame.pack(expand=True, fill="both", padx=20, pady=20)

            # Grid configuration for responsive layout
            main_frame.grid_columnconfigure((0,1), weight=1)
            main_frame.grid_rowconfigure(4, weight=1)

            # Get vehicle data
            
            # Fields and entries
            fields = [ "Marque", "Type", "Immatriculation","Service Utilisateur" , "Anne"
                    "Date Assurance", "Carburant","Date Control Technique " , "Conducteur"]
            entries = []
            
            
            # Create entries with improved layout
            for i,field in enumerate(fields):
                row = i // 2  # Integer division for row number
                col = i % 2   # Column number (0 or 1)
                
                # Container frame for each field
                field_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
                field_frame.grid(row=row, column=col, padx=10, pady=5, sticky="nsew")
                label = ctk.CTkLabel(field_frame, text=field + ":", font=("Helvetica", 18))
                label.pack(anchor="w", pady=(0, 2))
                
            
                
                
                    
                    
                    
               
                entry = ctk.CTkEntry(field_frame, font=("Helvetica", 13),placeholder_text=field, width=220,border_color="", height=40)
                entry.pack(fill="x")
                
                entries.append(entry)
                
            
            
            

            # Button frame
            button_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
            button_frame.grid(row=5, column=0, columnspan=2, pady=20, sticky="e")

            def save_record(): 
                
                
                
                values = [entry.get().strip() for entry in entries]
                
                if any(value == '' for value in [values[1], values[2]]):
                    messagebox.showerror("Error", "Please fill all fields")
                    return
                
                query = "INSERT INTO {tab} (marque, type, Immatriculation, [service_utilisateur],Anne, [date_assurance], carburant,[date_control_technique] , conducteur) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)"
                try:
                    connection = get_connection()
                    cursor = connection.cursor()
                    cursor.execute(query, values)
                    connection.commit()
                    fetch_all_data(tree, tab)
                    messagebox.showinfo("Success", "Record added successfully")
                    popup.destroy()
                except pyodbc.Error as e:
                    messagebox.showerror("Database Error", f"Failed to add record:{str(e)}")

            # Create buttons
            ctk.CTkButton(button_frame, text="Cancel", width=100, height=30, command=popup.destroy).pack(side="right", padx=5)
            ctk.CTkButton(button_frame, text="Clear", width=100, height=30, command=lambda: [entry.delete(0, 'end') for entry in entries]).pack(side="right", padx=5)
            ctk.CTkButton(button_frame, text="Save", width=100, height=30, command=save_record).pack(side="right", padx=5)

            # Center popup
            popup.update_idletasks()
            x = (popup.winfo_screenwidth() - popup.winfo_width()) // 2
            y = (popup.winfo_screenheight() - popup.winfo_height()) // 2
            popup.geometry(f"+{x}+{y}")
        
        def fetch_data(tree, query, params=()):
            try:
                connection = get_connection()
                cursor = connection.cursor()
                cursor.execute(query, params)
                data = cursor.fetchall()
                
                current_ids = {str(row[0]) for row in data}

                # Remove buttons for rows that no longer exist
                for vehicule_id in list(self.existing_button_frames.keys()):
                    if vehicule_id not in current_ids:
                        self.existing_button_frames[vehicule_id].destroy()
                        del self.existing_button_frames[vehicule_id]
                # Clear Treeview before inserting new data
                for item in tree.get_children():
                    tree.delete(item)
                
                # Insert fetched data into the Treeview
                for i, row in enumerate(data):
                    tag = 'evenrow' if i % 2 == 0 else 'oddrow'
                    vehicule_id=row[0]
                    tree.insert('', 'end', values=[str(item) for item in row], tags=(tag,))
                    if vehicule_id not in self.existing_button_frames:
                        add_row_buttons(tree, vehicule_id)
                

                cursor.close()
                connection.close()
            except pyodbc.Error as e:
                print(f"Error: {e}")
        
        def add_row_buttons(tree, vehicule_id):
            """Dynamically add Update, Delete, and Inspect buttons next to each row."""
            if vehicule_id in self.existing_button_frames:
                return
            # Find the corresponding row in the Treeview
            row_id = None
            for item in tree.get_children():
                values = tree.item(item)['values']
                if values and str(values[0]) == str(vehicule_id):
                    row_id = item
                    break

            
            # Create a frame to hold the buttons
            row_button_frame = ctk.CTkFrame(self.right_buttons_frame, fg_color="transparent",height=40)
            row_button_frame.pack(fill="x",pady=4)  # Expand to fill the row

            # Create buttons inside this frame
            update_btn = ctk.CTkButton(row_button_frame, text="", fg_color="transparent",width=30,image=self.Update_icon,
                                    command=lambda: update_popup(tree, tab,row_id))
            delete_btn = ctk.CTkButton(row_button_frame, text="", fg_color="transparent",width=30,image=self.supprimer_icon ,
                                    command=lambda: delete_selected(tree, tab,row_id))
            inspect_btn = ctk.CTkButton(row_button_frame, text="", fg_color="transparent",width=30,image=self.Inspect_icon,command=lambda: show_inspect_window(self, tree, row_id) )
            if self.user_role == "technicien":
                update_btn.configure(state="disabled", fg_color="gray")
                delete_btn.configure(state="disabled", fg_color="gray")
            # Pack buttons inside the frame
            update_btn.pack(side="left", padx=2)
            delete_btn.pack(side="left", padx=2)
            inspect_btn.pack(side="left", padx=2)
            
            self.existing_button_frames[vehicule_id] = row_button_frame
        
       
        
        
            
            
            
        def filter_search(tree,*args):
                
                
            
            value = self.search_entry.get()
            query = f'''SELECT  vehicule_id,marque, type, Immatriculation,service_utilisateur,conducteur  
                    FROM Vehicule WHERE 
                    vehicule_id LIKE ? OR 
                    marque LIKE ? OR 
                    type LIKE ? OR 
                    Immatriculation LIKE ? OR 
                    service_utilisateur LIKE ? OR 
                    conducteur  LIKE ?'''
            params = (f"%{value}%",) * 6 
            fetch_data(tree, query, params)
            
                
            
                            
                
# Fetch all data
        def fetch_all_data(tree, tab):
            
            query = """
             SELECT  vehicule_id,marque, type, Immatriculation,service_utilisateur,conducteur  
                    FROM Vehicule 
                """
            fetch_data(tree, query)
            if hasattr(self, "main_app") and self.main_app is not None:
                if hasattr(self.main_app, "highlighted_rows"):
                    print(" main_app.highlighted_rows is accessible")
                today = datetime.date.today()
                
                connection = get_connection()
                cursor = connection.cursor()
                for item in tree.get_children():
                    values = tree.item(item)['values']
                    vehicule_id = values[0]  # Assuming first column is ID
                    
                    if vehicule_id in self.main_app.highlighted_rows:
                        cursor.execute(
                            "SELECT date_assurance, date_control_technique "
                            "FROM Vehicule WHERE vehicule_id = ?",
                            (vehicule_id,)
                        )
                        assurance_date, control_date = cursor.fetchone()
                        if not assurance_date or not control_date:
                            continue
                        # Check if dates are still expired
                        if assurance_date <= today or control_date <= today:
                            for vehicule_id in self.main_app.highlighted_rows:
                                for item in tree.get_children():
                                    values = tree.item(item)['values']
                                    if str(values[0]) == str(vehicule_id):
                                        tree.item(item, tags=("highlighted",))
                        else:
                            # Remove from tracking if dates are fixed
                            del self.main_app.highlighted_rows[vehicule_id]
                            tree.item(item, tags=("normalrow",))
                
                    print(f" Error: vehicule_id {vehicule_id} not found in highlighted_dates")  
                connection.close()
        
        def delete_selected( tree, tab,row_id):
            # Delete the selected row from the database
            if not messagebox.askyesno("Confirm Delete", "Delete This record?"):
                return

            try:
                connection = get_connection()
                cursor = connection.cursor()
                
                # Delete all selected records
                current_values = tree.item(row_id)['values']
                
                record_id = current_values[0]  # Assuming first column is ID
                cursor.execute(f"DELETE FROM {tab} WHERE vehicule_id = ?", (record_id,))
                
                # Remove associated button frame
                if record_id in self.existing_button_frames:
                    self.existing_button_frames[record_id].destroy()
                    del self.existing_button_frames[record_id]

                connection.commit()
                messagebox.showinfo("Success", f"Deleted {row_id} record")
                
                # Refresh data
                fetch_all_data(tree, tab)

            except pyodbc.Error as e:
                messagebox.showerror("Database Error", f"Failed to delete records: {str(e)}")
            finally:
                if 'connection' in locals():
                    connection.close()
        def update_selected(tree, tab, entries,row_id):
         

            item_values = tree.item(row_id)['values']
            if not item_values:
                return

            record_id = item_values[0]  # Assuming the first column is the unique ID
            new_values = [entries[3].get().strip(),entries[4].get().strip(),entries[5].get().strip(),entries[6].get().strip(),entries[7].get().strip()]
            if any(value == '' for value in new_values):
                messagebox.showwarning("Warning", "Please fill all fields")
                return

            query = f"UPDATE {tab} SET  {tab} SET marque = ?, type = ?, Immatriculation = ?, service_utilisateur = ?,Anne = ?,[date_assurance] = ?, carburant = ?,[date_control_technique] = ?, conducteur = ? WHERE vehicule_id = ?"
            try:
                connection = get_connection()
                cursor = connection.cursor()
                cursor.execute(query, (*new_values, record_id))
                connection.commit()

                # Clear entry fields after successful update
                for entry in entries:
                    entry.delete(0, tk.END)

                # Refresh treeview
                fetch_all_data(tree, tab)
                print("Record updated successfully.")
            except pyodbc.Error as e:
                print(f"Error: {e}")
        def update_popup(tree, tab,row_id):
            current_values = tree.item(row_id)['values']
            glac_num = current_values[0]
            # Fetch complete data using existing pattern
            query = """SELECT  marque, type, Immatriculation,service_utilisateur,[Anne],[date_assurance],[carburant],[date_control_technique],conducteur  
                    FROM Vehicule
                    WHERE vehicule_id = ?"""
            
            try:
                connection = get_connection()
                cursor = connection.cursor()
                cursor.execute(query, (glac_num,))
                full_data = cursor.fetchone()
            except pyodbc.Error as e:
                messagebox.showerror("Database Error", f"Failed to fetch data: {str(e)}")
                return
            finally:
                if connection:
                    connection.close()
            print("THIS THE ROW ID ",row_id)
            # Create popup window
            popup = ctk.CTkToplevel()
            popup.title("Update Record")
            popup.geometry("600x450")
            popup.resizable(False, False)
            popup.transient(tree.winfo_toplevel())  # Make popup transient to main window
            popup.grab_set()  # Make popup modal
            popup.focus_set()
            
            # Create main frame with padding
            main_frame = ctk.CTkFrame(popup)
            main_frame.pack(expand=True, fill="both", padx=20, pady=20)
    
            # Grid configuration for responsive layout
            main_frame.grid_columnconfigure((0,1), weight=1)
            main_frame.grid_rowconfigure(4, weight=1)

            # Fields and entries
            fields = ["N°","Marque", "Type", "Immatriculation","Service Utilisateur" , "Anne"
                    "Date Assurance", "Carburant","Date Control Technique " , "Conducteur"]
            entries = []

            # Create entries with improved layout
            for i, field in enumerate(fields):
                row = i // 2  # Integer division for row number
                col = i % 2   # Column number (0 or 1)
                
                # Container frame for each field
                field_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
                field_frame.grid(row=row, column=col, padx=10, pady=5, sticky="nsew")
                
                label = ctk.CTkLabel(field_frame, text=field + ":", font=("poppins", 18))
                label.pack(anchor="w", pady=(0, 2))
                
                entry = ctk.CTkEntry(field_frame,placeholder_text=field,border_color="", font=("poppins", 13), width=220,height=40)
                entry.pack(fill="x")
                
                entries.append(entry)
               
            # Populate entries with current values
            for entry, value in zip(entries, full_data):
                
                entry.delete(0, 'end')
                entry.insert(0, str(value))
                

            # Button frame at bottom right
            button_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
            button_frame.grid(row=6, column=0, columnspan=2, padx=10, pady=10, sticky="e")

            def confirm_update():
                update_selected(tree, tab, entries,row_id)
                popup.destroy()

            # Create buttons
            cancel_btn = ctk.CTkButton(
                button_frame, 
                text="Annuler",
                font=("Helvetica", 11),
                width=100,
                height=30,
                command=popup.destroy
            )
            cancel_btn.pack(side="right", padx=10)

            confirm_btn = ctk.CTkButton(
                button_frame, 
                text="Affecter",
                font=("Helvetica", 11),
                width=100,
                height=30,
                command=confirm_update
            )
            confirm_btn.pack(side="right")

            # Center the popup on screen
            popup.update_idletasks()
            x = (self.winfo_screenwidth() - popup.winfo_width()) // 2
            y = (self.winfo_screenheight() - popup.winfo_height()) // 2
            popup.geometry(f"+{x}+{y}")
        def show_inspect_window(self, tree, row_id):
            current_values = tree.item(row_id)['values']
            veh_id = current_values[0]  # Get the glaciol number

            # Fetch complete data including new fields
            try:
                connection = get_connection()
                cursor = connection.cursor()
                cursor.execute("""
                    SELECT  vehicule_id,marque, type, Immatriculation,service_utilisateur,[Anne],[date_assurance],[date_control_technique],[carburant],conducteur  
                    FROM Vehicule
                    WHERE vehicule_id = ?
                """, (veh_id,))
                full_data = cursor.fetchone()
            except pyodbc.Error as e:
                messagebox.showerror("Database Error", f"Failed to fetch details: {str(e)}")
                return
            finally:
                if connection:
                    connection.close()
            # Get all items and current index
            all_items = tree.get_children()
            current_index = all_items.index(row_id) 
            
            # Create sliding window
            inspect_win = ctk.CTkToplevel()
            inspect_win.withdraw()
            inspect_win.overrideredirect(True)
            inspect_win.attributes("-topmost", True)
            inspect_win.fg_color = ("#2b2b2b", "#2b2b2b")
            
            window_width = 300  # Set your desired width
            window_height = 700  # Set your desired height
            screen_width = self.winfo_screenwidth()
            screen_height = self.winfo_screenheight()
                    
            # Animation variables
            target_x = screen_width - window_width  # 20px margin from right
            current_x = screen_width + 500  # Start fully off-screen
            current_y = (screen_height - window_height) // 2 
            fade_alpha = 0.0

            
            
            # Navigation variables
            main_container = ctk.CTkFrame(inspect_win, fg_color="#333333",corner_radius=20)
            main_container.pack(fill="both", expand=True, padx=10, pady=10)

            # Close button (fixed position)
            close_btn = ctk.CTkButton(
                main_container,
                text="✕",
                command=lambda: fade_out(),
                fg_color="transparent",
                width=30,
                height=30,
                font=("Arial", 16)
            )
            close_btn.place(relx=0.98, rely=0.02, anchor="ne")

            # Content frame (scrollable)
            content_frame = ctk.CTkScrollableFrame(main_container, fg_color="transparent")
            content_frame.pack(fill="both", expand=True, pady=40)
            
            inspect_win.geometry(f"{window_width}x{window_height}+{current_x}+{current_y}")
            inspect_win.deiconify()
            inspect_win.attributes("-alpha", 0.01)
            
            def update_display():
                nonlocal current_index
                item_id = all_items[current_index]
                values = tree.item(item_id, 'values')
                fields = [
                        ("N°", full_data[0]),
                        ("Marque", full_data[1]),
                        ("Type", full_data[2]),
                        ("Immatriculation", full_data[3]),
                        ("Service Utilisateur", full_data[4]),
                        ("Anne", full_data[5]),
                        ("Date d'assurance", full_data[6]),  
                        ("Date de controle Techique", full_data[7]),
                        ("Carburant",full_data[8]),
                        ("Conducteur",full_data[9])
                        ]

                
                # Clear previous entries
                for widget in content_frame.winfo_children():
                    widget.destroy()
                
                # Display fields
                for field_name, value in fields:
                    row_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
                    row_frame.pack(fill="x", pady=5)
                    
                    label1 = ctk.CTkLabel(row_frame, 
                                        text=f"{field_name}:",
                                        font=("poppins", 12, "bold"))
                    label1.pack(anchor="w")
                    
                    entry1 = ctk.CTkEntry(row_frame,height=40,width=250, font=("poppins", 12))
                    entry1.insert(0, str(value))
                    entry1.configure(state="readonly")
                    entry1.pack(side="left",pady=5)

                    # Second column if exists
                    
                
                # Navigation buttons
                btn_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
                btn_frame.pack(side="bottom", fill="x", pady=20)
                
                prev_btn = ctk.CTkButton(
                    btn_frame, 
                    text="◄ Previous",width=80,
                    fg_color="#534AE1" if current_index > 0 else "gray30",
                    command=lambda: navigate(-1),
                    state="normal" if current_index > 0 else "disabled"
                )
                prev_btn.pack(side="left", padx=10)
                
                next_btn = ctk.CTkButton(
                    btn_frame, 
                    text="Next ►",width=80,
                    fg_color="#534AE1" if current_index < len(all_items)-1 else "gray30",
                    command=lambda: navigate(1),
                    state="normal" if current_index < len(all_items)-1 else "disabled"
                )
                next_btn.pack(side="right", padx=10)
                
                # Update treeview selection
                tree.selection_set(all_items[current_index])
                tree.see(all_items[current_index])
            
            def navigate(direction):
                nonlocal current_index
                new_index = current_index + direction
                if 0 <= new_index < len(all_items):
                    current_index = new_index
                    update_display()

            def fade_out():
                def animate():
                        nonlocal current_x
                        current_x += 30  # Move right while fading
                        inspect_win.geometry(f"{window_width}x{window_height}+{current_x}+{current_y}")
                        
                        nonlocal fade_alpha
                        fade_alpha = max(0, fade_alpha - 0.1)
                        inspect_win.attributes("-alpha", fade_alpha)
                        
                        if fade_alpha > 0:
                            inspect_win.after(20, animate)
                        else:
                            inspect_win.destroy()
                animate()

            # Bind window close events
            def check_click(event):
                x, y = event.x_root, event.y_root
                win_x = inspect_win.winfo_x()
                win_y = inspect_win.winfo_y()
                win_width = inspect_win.winfo_width()
                win_height = inspect_win.winfo_height()
                
                if not (win_x <= x <= win_x + win_width and
                        win_y <= y <= win_y + win_height):
                    fade_out()
            inspect_win.bind("<Button-1>", check_click)
            # Start animations
            def slide_in():
                nonlocal current_x, fade_alpha
                if current_x > target_x:
                    current_x -= 30
                    fade_alpha = min(fade_alpha + 0.05, 1.0)  # Gradually increase opacity
                    inspect_win.attributes("-alpha", fade_alpha)
                    inspect_win.geometry(f"{window_width}x{window_height}+{current_x}+{current_y}")
                    inspect_win.after(15, slide_in)
                else:
                    inspect_win.attributes("-alpha", 1) 

            update_display()
            slide_in()
        def export_visible_to_excel():



            # Collect all visible num_glaciol IDs from the Treeview
            num_glaciol_list = []
            for item in tree.get_children():
                row_values = tree.item(item)['values']
                if row_values:
                    num_glaciol_list.append(row_values[0])  # Assuming col1 is num_glaciol (primary key)

            if not num_glaciol_list:
                messagebox.showwarning("No Data", "No data to export!")
                return
            print(num_glaciol_list)
            # Prepare SQL query to fetch full data of only visible rows
            id_list = ','.join(str(glac_id) for glac_id in num_glaciol_list)
            print(id_list)
            query = f"""
                SELECT vehicule_id AS [N°], marque AS [Marque], type AS [Type], Immatriculation AS [Immatriculation], 
                    service_utilisateur AS [Service Utilisateur], Anne AS [Anne], date_assurance AS [Date d'assurance], date_control_technique AS [Date de controle Techique],carburant AS [Carburant],conducteur AS [Conducteur]
                FROM Vehicule 
                
                WHERE vehicule_id IN ({id_list})
            """

            try:
                connection = get_connection()
                cursor = connection.cursor()
                cursor.execute(query)
                rows = cursor.fetchall()
                print(rows)
                # Column names from SQL aliases (inspect view)
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Glaciol Export"

                # Write headers (as in inspect window / database)
                headers = ["N°","Marque", "Type", "Immatriculation","Service Utilisateur" , "Anne"
                    "Date d'Assurance", "Carburant","Date de Control Technique " , "Conducteur"]
                ws.append(headers)

                # Write data rows manually
                for db_row in rows:
                    ws.append(list(db_row))  # Ensure it's a list
                header_font = Font(bold=True, color="FFFFFF", size=12)
                header_fill = PatternFill(fill_type="solid", fgColor="4F81BD")  # Blue background
                alignment = Alignment(horizontal="left", vertical="center")
                ws.row_dimensions[1].height = 30
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="left", vertical="center")    
                #   Adjust Column Widths
                for col_num, header in enumerate(headers, 1):
                    max_length = max((len(str(ws.cell(row=row, column=col_num).value)) for row in range(1, ws.max_row + 1)), default=10)
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width
                # Save the workbook
                file_path = os.path.abspath("Vehicules_Export.xlsx")
                wb.save(file_path)

                messagebox.showinfo("Export Successful", f"Data exported to {file_path}")

                # Auto-open the Excel file
                
                os.startfile(file_path)  # Windows
              
            except Exception as e:
                messagebox.showerror("Export Error", str(e))
            finally:
                if 'connection' in locals():
                    connection.close()

            
        tab = "Vehicule "
        # Style
        style = ttk.Style()
        style.theme_use("classic")
        style.configure("Treeview", background="#050505", foreground="#b3b3b3",rowheight=35, fieldbackground="#050505",
        borderwidth=0,
        font=('Poppins', 12))
        style.configure("Treeview.Heading",  background="#171717",rowheight=50, foreground="#cccccc",relief="flat",borderwidth=0,
    font=('Poppins', 14, 'bold'),
    padding=5)
        
        
        # Add alternating row colors
        style.map("Treeview", 
        background=[('selected', '#455e88')],  # Blue background when selected
        foreground=[('selected', '#cccccc')],
        font=[('selected',('Poppins', 12, 'bold'))] )       
# Add this updated style configuration before creating the Treeview

    
    
# Hover and selection effects
        

        style.map("Treeview.Heading",
    background=[('active', '#333333')],
    relief=[('pressed', 'groove'), ('!pressed', 'groove')]
    )



# Modern scrollbar styling
        style.configure("Treeview.Scrollbar",
    troughcolor="#f0f0f0",
    background="#c1c1c1",
    borderwidth=0,
    relief="flat"
)


# Then modify your Treeview creation to use the custom style:
        
        tree = ttk.Treeview(
    self.page_frame,
    columns=("col1", "col2", "col3", "col4", "col5", "col6"),
    show='headings',
    style="Treeview",
    selectmode="extended"
    )
        self.tree = tree
        sort_direction = {
        "col1": None,
        "col2": None, 
        "col3": None,
        "col4": None,
        "col5": None,
        "col6": None,
        
    }   
        column_headings = {
    "col1": "N°",
    "col2": "Marque",
    "col3": "Type", 
    "col4": "Immatriculation",
    "col5": "Service Utilisateur",
    "col6": "Conducteur"
        
    }
        ord_column_headings = {
    "col1": "vehicule_id",
    "col2": "Marque",
    "col3": "Type", 
    "col4": "Immatriculation",
    "col5": "service_utilisateur",
    "col6": "Conducteur",
    }
    
        def treeview_sort_column(tree, col, tab):
    # Get the field name corresponding to the column
            field = ord_column_headings[col]
    
    # Toggle sort direction
            if sort_direction[col] is None or sort_direction[col] == 'DESC':
               sort_direction[col] = 'ASC'
            else:
               sort_direction[col] = 'DESC' 
    
    # Construct and execute SQL query with ORDER BY
            query = f"""
            SELECT  vehicule_id,marque, type, Immatriculation,service_utilisateur,conducteur  
                    FROM Vehicule 
            ORDER BY {field} {sort_direction[col]}"""
            fetch_data(tree, query)
    
# Set column headings and properties
        for col, heading in column_headings.items():
            tree.heading(col, text=heading, anchor=tk.W,command=lambda c=col: treeview_sort_column(tree, c, tab))
            if col == "col1" :
                tree.column(col, anchor=tk.W, width=40, minwidth=60)
            else:
                tree.column(col, anchor=tk.W, width=180, minwidth=100)

 

# Configure horizontal scrollbar style
       
# Add these lines for alternating row colors
        tree.tag_configure('oddrow', background='#08090b')
        tree.tag_configure('evenrow', background='#050505')


      

        

# Configure the tree to use the scrollbars
        

# Grid layout for scrollbars
        tree.grid(row=1, column=0,  padx=(20,3), pady=5, sticky="nsew")
       
        # Fetch initial data
        fetch_all_data(tree, tab)
        
        self.search_entry=ctk.CTkEntry(self.buttons_frame,corner_radius=20,border_width=1,border_color="",placeholder_text="Search")
        self.search_entry.grid(row=0, column=0, pady=10,padx=(0,320),sticky="w",ipadx=150)
        self.search_entry.bind('<FocusIn>', lambda e, entry=self.search_entry: entry.configure(border_color="#561B8D"))
        self.search_entry.bind('<FocusOut>', lambda e, entry=self.search_entry: entry.configure(border_color=""))
        self.search_entry.bind('<KeyRelease>', lambda e: filter_search(tree))
        
        self.add_button = ctk.CTkButton(self.buttons_frame,width=25,height=25,text="Ajouter",image=self.add_icon,fg_color="#534ae1",corner_radius=30,compound="left",command=lambda: start_add_mode(tree, tab, ))
        self.add_button.grid(row=0, column=1, padx=10, pady=10,sticky="e")
        if self.user_role == "technicien":
            self.add_button.configure(state="disabled", fg_color="gray")
        self.export_button = ctk.CTkButton(
                self.buttons_frame,
                text="Export to Excel",
                fg_color="#4CAF50",  # Green button
                corner_radius=30,
                command=export_visible_to_excel
            )
        self.export_button.grid(row=0, column=3, padx=10, pady=10)
        self.filter_button = ctk.CTkButton(self.buttons_frame,width=25,height=25,text="Filter",image=self.filter_icon,compound="left",fg_color="transparent",corner_radius=30, command=lambda: filter_search(tree, tab))
        self.filter_button.grid(row=0, column=4, padx=10, pady=10)

        self.refresh_button = ctk.CTkButton(self.buttons_frame,width=30 ,text="Refresh",image=self.refresh_icon,compound="left",fg_color="transparent",corner_radius=30,command=lambda: fetch_all_data(tree, tab))
        self.refresh_button.grid(row=0, column=2, padx=10, pady=10)
        
        
        
         
        
        
        






'''

if __name__ == "__main__":
    app = MainPage()

    app.mainloop()'''
